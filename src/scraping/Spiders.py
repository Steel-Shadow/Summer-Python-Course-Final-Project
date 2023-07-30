import multiprocessing
import os
import time
from typing import Iterable

import openpyxl

from websites import icourses, bilibili, netease, cmooc, tencent, haodaxue, imooc, mooc


class Spiders:
    # 网页 -> 抓取函数
    web_scraping = {
        'icourses': icourses.scraping,
        'bilibili': bilibili.scraping,
        'netease': netease.scraping,
        'cmooc': cmooc.scraping,
        'tencent': tencent.scraping,
        'haodaxue': haodaxue.scraping,
        'imooc': imooc.scraping,
        'mooc': mooc.scraping,
    }

    @staticmethod
    def err_call_back(err):
        print(f'多进程爬取error：{str(err)}')

    @staticmethod
    def task(shared_dict, web: str, keywords: Iterable[str]):
        scraping = Spiders.web_scraping.get(web)
        gen_scraping = scraping(keywords)

        # yield迭代返回 每次返回单个 keyword 的爬取结果
        iter_keywords = iter(keywords)
        for iter_scraping in gen_scraping:
            shared_dict[(web, next(iter_keywords))] = iter_scraping

    @classmethod
    def scraping(cls, keywords: Iterable[str]) -> list:
        """
        请尽量少调用(把关键词放一块)，每次抓取都要重新打开、关闭浏览器
        """
        pool = multiprocessing.Pool()
        manager = multiprocessing.Manager()

        # 网页 -> 抓取信息 多进程共享
        shared_dict = manager.dict()

        for web in cls.web_scraping:
            # 爬取任务加入进程池
            pool.apply_async(cls.task, (shared_dict, web, keywords), error_callback=cls.err_call_back)

        pool.close()
        pool.join()

        lst = list()
        for i in shared_dict.values():
            if i:
                lst.append(i)
        return lst

    @classmethod
    def scraping_dict(cls, keywords: Iterable[str]) -> dict:
        pool = multiprocessing.Pool()
        manager = multiprocessing.Manager()

        shared_dict = manager.dict()

        for web in cls.web_scraping:
            # 爬取任务加入进程池
            pool.apply_async(cls.task, (shared_dict, web, keywords), error_callback=cls.err_call_back)

        pool.close()
        pool.join()
        return shared_dict


if __name__ == '__main__':
    """
    修改words为要爬取的关键词，运行完毕后生成 scraping_offline.xlsx
    这里根据CS_KG.txt，生成xlsx
    """
    with open(os.path.join(os.path.dirname(__file__), 'CS_KG.txt'), encoding='utf-8') as in_file:
        graph = in_file.readlines()

    words = list()

    for line in graph:
        words.append(line.strip())
    # words = ['系统']

    begin = time.time()
    res = Spiders.scraping_dict(words)
    end = time.time()
    print(end - begin)

    # 创建一个新的工作簿
    workbook = openpyxl.Workbook()

    # 选择默认的活动工作表
    sheet = workbook.active

    for row in range(1, len(words) + 1):
        # 第一列 关键词
        keyword = words[row - 1]
        sheet.cell(row=row, column=1, value=keyword)
        column = 1
        for w in Spiders.web_scraping:  # 网页
            courses = res.get((w, keyword))

            if not courses:
                continue

            for course in courses:  # 课程
                unit = ''
                for info in course:  # 课程各个信息
                    unit += info + ','
                column += 1
                sheet.cell(row, column, unit)

    # 保存工作簿
    workbook.save("scraping_offline.xlsx")

    # 关闭工作簿
    workbook.close()
